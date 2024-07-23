""" Processing data in the "Table Agent*.xlsx" file specified as the 
    first command line argument and load data into "Table Agent Summary.xlsx"

    usage:  get_summary_app_id.py  "Tenable*.xlsx" "Hardware*.xlsx"

    output: app_id_xxxx_summary.xlsx
   
"""
import sys
import re
from openpyxl import load_workbook, Workbook
from datetime import datetime


# database issue path patten
# /oracle/crs/19.3.0.0.0/grid/jdk/
# /oracle/crs/19.3.0.0.0/grid/OPatch_bkp/jre/

def  is_db_path_issue(path):
    r = re.compile(r'\boracle|\bopatch|\bCassandr|SQL\sServer|MySQL|PostgreSQL|\bsplex|\bggate|goldengate|\bgrid|\bOEM|\bdse-', flags=re.I | re.X)
    if r.search(path):
       print("Path matching: {}".format(path))
       return True
    return False


def add_path_to_dict(text, pathDict, vulnName, host):
    p_list = re.findall(r'Path\s*:\s*(.+?)\s*\n', text)
    for p in [ p.strip() for p in p_list ] :
        # check if p match database issue patthen
        if is_db_path_issue(p):
            if p in pathDict.keys(): 
                pathDict[p].append(host)
            else :     
                pathDict[p] =[]
                pathDict[p].append(vulnName)
                pathDict[p].append(host)
    

def print_rows(sheet):
    for row in sheet.iter_rows(values_only=True):
        print(row)

if __name__ == "__main__":

    if len(sys.argv) < 3 :
        print("Not enough arguments")
        print('''Usage:  get_summary_vast_id.py  "Tenable*.xlsx" "Hardware*.xlsx"''')
        exit(1)

    vulnSheetFile=sys.argv[1]
    hardSheetFile=sys.argv[2]

    # create the output excel
    m = re.search(r'app_id_\d+', vulnSheetFile)
    app_id_str = m.group()
    outputfile = app_id_str +'_summary.xlsx'
    
    workbook_out = Workbook()

    sheet_path=workbook_out.active
    sheet_path.title = "path"
    sheet_path["A1"] = "Path Name"
    sheet_path["B1"] = "Vuln Name"
    sheet_path["C1"] = "Server List"

    workbook_out.create_sheet('Vulns')
    sheet_vulns =workbook_out['Vulns']
    sheet_vulns["A1"] = "Vulns Name"
    sheet_vulns["B1"] = "Summary/Description"
    sheet_vulns["C1"] = "Server list"

    issueDict={}

    # hardware sheet
    print('\nList DB related vulnerabilities in "{}"\n\n'.format(hardSheetFile))
    wb = load_workbook(filename = hardSheetFile)
    ws = wb['Vulnerabilities']

    row_count = ws.max_row

    # Vuln_name => [desc, host1, host2 ...]
    # L  

    for i in range(row_count):
        risk = ws.cell(row=i+1, column=7).value
        role = ws.cell(row=i+1, column=12).value
        if role == "Database" and ( risk == "critical" or risk == "medium" or risk =="high" ) :
            vulnName = ws.cell(row=i+1, column=12).value + ' (Hardware)'
            summary = ws.cell(row=i+1, column=13).value
            host = ws.cell(row=i+1, column=2).value
            if vulnName in issueDict.keys(): 
                issueDict[vulnName].append(host)
            else: 
                issueDict[vulnName] =[]
                issueDict[vulnName].append(summary)
                issueDict[vulnName].append(host)

    # Tenable sheet 
    print('\nList DB related vulnerabilities in "{}"\n\n'.format(vulnSheetFile))
    wb = load_workbook(filename = vulnSheetFile)
    ws = wb.active
    row_count = ws.max_row
    pathDict={}

    #
    # Column W = 23  vuln name
    # Column X = 24  Family
    # Column Y = 25  Desc 
    #

    for i in range(row_count):
        family = ws.cell(row=i+1, column=24).value
        host = ws.cell(row=i+1, column=2).value
        vulnName = ws.cell(row=i+1, column=23).value
        if family == 'Databases' :
            vulnName = ws.cell(row=i+1, column=23).value
            issueDesc = ws.cell(row=i+1, column=25).value
            if vulnName in issueDict.keys(): 
                issueDict[vulnName].append(host)
            else: 
                issueDict[vulnName] =[]
                issueDict[vulnName].append(issueDesc)
                issueDict[vulnName].append(host)
        elif family == 'Misc.' : 
            Plugin_Output = ws.cell(row=i+1, column=31).value
            add_path_to_dict(Plugin_Output, pathDict, vulnName, host)
    

    # write vulns issue data to outputfile

    for i in issueDict:
        print("Vuln issue   : {} Adding ...".format(i))
        sheet_vulns.insert_rows(idx=2)
        sheet_vulns["A2"] = i
        issueList = issueDict[i]
        sheet_vulns["B2"] = issueList[0] 
        for s in issueList[1:]:
            if sheet_vulns["C2"].value:
                sheet_vulns["C2"] =  sheet_vulns["C2"].value + "," + s  
            else:
                sheet_vulns["C2"] = s   


    # write path data to outputfile

    for p in pathDict: 
        # print("### Processing - {} Adding ...".format(p))
        sheet_path.insert_rows(idx=2)
        sheet_path["A2"] = p 
        serverList = pathDict[p] 
        sheet_path["B2"] = serverList[0] 
        for s in serverList[1:]:
            if sheet_path["C2"].value:
                sheet_path["C2"] =  sheet_path["C2"].value + "," + s  
            else:
                sheet_path["C2"] = s   


    workbook_out.save(filename=outputfile)