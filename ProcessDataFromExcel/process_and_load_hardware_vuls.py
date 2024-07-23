"""Processing data in a file called "Hardware_Vast*.xlsx specified as the first
   command line argument and load the data into Hardware summary.xlsx table

   note:
   
    The purpose is to generate a list of unique vulnerabilities from the input files and save it to the output file nameed sheets\Hardware Vulns Summary.xlsx
"""
import sys
from openpyxl import load_workbook
from datetime import datetime

if __name__ == "__main__":

    sheetFile=sys.argv[1]

    print('\nList DB related vulnerabilities in "{}"\n\n'.format(sheetFile))

    wb = load_workbook(filename = sheetFile)
    ws = wb['Vulnerabilities']

    row_count = ws.max_row

    # key=>value : Vuln_name => [desc, host1, host2 ...]
    issueDict={}
    
    # Column L = 12  vuln name
    # Column M = 13  Summary 
    # Column B = 2   Host 

    for i in range(row_count):
        # col 7 - risk
        risk = ws.cell(row=i+1, column=7).value
        if risk == "critical" or risk == "medium" :
            vulnName = ws.cell(row=i+1, column=12).value
            summary = ws.cell(row=i+1, column=13).value
            host = ws.cell(row=i+1, column=2).value
            if vulnName in issueDict.keys(): 
                issueDict[vulnName].append(host)
            else: 
                issueDict[vulnName] =[]
                issueDict[vulnName].append(summary)
                issueDict[vulnName].append(host)
    
        
    # load into an excel
    wb = load_workbook(filename = "sheets\Hardware Vulns Summary.xlsx")
    ws = wb["Vuln"]
    
    colA = ws["A"]
    colA_values =[c.value for c in colA]

    # print(colA_values)
    
    for i in issueDict:
        # print(i)
        if i  in  colA_values :
            print("### INFO    - {} exists".format(i))
        else :
            print("### Warning - {} is not found. Adding ...".format(i))
            ws.insert_rows(idx=2)
            ws["A2"] = i
            issueList = issueDict[i]
            ws["B2"] = issueList[0] 

    ws_sheets = wb["Sheets"]
    ws_sheets.insert_rows(idx=2)
    ws_sheets["A2"] = sheetFile 

    now = datetime.now() # current date and time
    date_time = now.strftime("%m/%d/%Y, %H:%M:%S")
    ws_sheets["B2"] = date_time 

    wb.save(filename = "sheets\Hardware Vulns Summary.xlsx")
